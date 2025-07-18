import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from excel_report.chart_builder import add_price_chart

def generate_excel_report(data: list, filename="samples/output_examples.xlsx"):
    """
    Generate an Excel report from product data.
    Includes filters and summary chart.
    """
    if not data:
        print("⚠️ No data provided to generate report.")
        return

    df = pd.DataFrame(data)
    df.sort_values(by="price", inplace=True)

    # Save to Excel
    df.to_excel(filename, index=False)

    # Add filters and charts
    wb = load_workbook(filename)
    ws = wb.active

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws["A2"]  # Freeze header row

    # Add charts using helper
    add_price_chart(wb, ws, df)

    wb.save(filename)
    print(f" Excel report saved to: {filename}")
